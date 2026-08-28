"""Employee management: scoped list + CRUD + Excel import/export."""
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel

from . import audit, auth, commands, db

router = APIRouter(prefix="/api/employees", tags=["employees"])
XLSX_MT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class EmployeeIn(BaseModel):
    pin: str | None = None            # required on create
    name: str | None = None
    first_name: str | None = None
    last_name: str | None = None
    card: str | None = None
    department_id: int | None = None
    position_id: int | None = None
    area_id: int | None = None
    gender: str | None = None
    mobile: str | None = None
    hire_date: str | None = None
    active: bool | None = None


SELECT = f"""
    SELECT e.pin, {db.FULLNAME} AS name, e.first_name, e.last_name, e.card, e.privilege_name,
           e.department_id, d.name AS department, e.position_id, p.name AS position,
           e.area_id, e.gender, e.mobile, e.hire_date,
           (e.photo IS NOT NULL AND e.photo<>'') AS has_photo, e.active, e.updated_at
    FROM employees e
    LEFT JOIN department d ON d.id = e.department_id
    LEFT JOIN position   p ON p.id = e.position_id
"""


def _scope_clause(user: dict):
    """Return (sql_fragment, params) restricting to the caller's departments."""
    ids = auth.scope_dept_ids(user)
    if ids is None:
        return "", []
    if not ids:
        return "AND FALSE", []
    return "AND e.department_id = ANY(%s)", [ids]


def _assert_can_touch_dept(user: dict, department_id: int | None):
    ids = auth.scope_dept_ids(user)
    if ids is None:
        return
    if department_id is None or department_id not in ids:
        raise HTTPException(403, "Outside your department scope")


@router.get("")
def list_employees(user: dict = Depends(auth.get_current_user),
                   search: str | None = None,
                   department_id: int | None = None,
                   active: bool | None = None,
                   page: int = 1, page_size: int = 25):
    scope_sql, params = _scope_clause(user)
    where = ["1=1"]
    if search:
        where.append(
            "(e.pin ILIKE %s OR e.name ILIKE %s OR e.first_name ILIKE %s "
            "OR e.last_name ILIKE %s "
            "OR TRIM(COALESCE(e.first_name,'')||' '||COALESCE(e.last_name,'')) ILIKE %s)")
        s = f"%{search}%"
        params += [s, s, s, s, s]
    if department_id is not None:
        where.append("e.department_id = %s")
        params.append(department_id)
    if active is not None:
        where.append("e.active = %s")
        params.append(active)
    where_sql = " AND ".join(where)
    total = db.query(
        f"SELECT count(*) AS c FROM employees e WHERE {where_sql} {scope_sql}",
        tuple(params))[0]["c"]
    page = max(1, page)
    off = (page - 1) * page_size
    rows = db.query(
        f"{SELECT} WHERE {where_sql} {scope_sql} ORDER BY e.pin "
        f"LIMIT %s OFFSET %s", tuple(params + [page_size, off]))
    return {"data": rows, "total": total, "page": page, "page_size": page_size}


@router.get("/inactive")
def inactive(months: int = 3, user: dict = Depends(auth.get_current_user)):
    """Active employees with no attendance in the last `months` months."""
    months = max(1, min(months, 24))
    scope_sql, params = _scope_clause(user)
    # "Inactive" = punched before but not in the last N months, OR an old record
    # that never punched. A freshly-added employee (no punches yet, created
    # recently) is NOT inactive — it just hasn't been used yet.
    rows = db.query(
        f"""SELECT e.pin, {db.FULLNAME} AS name, e.department_id, dep.name AS department,
                   (e.photo IS NOT NULL AND e.photo<>'') AS has_photo,
                   MAX(a.punch_time) AS last_punch
            FROM employees e
            LEFT JOIN attendance a ON a.pin = e.pin
            LEFT JOIN department dep ON dep.id = e.department_id
            WHERE e.active {scope_sql}
            GROUP BY e.pin, e.first_name, e.last_name, e.name, e.department_id,
                     dep.name, e.photo, e.created_at
            HAVING (MAX(a.punch_time) IS NOT NULL
                    AND MAX(a.punch_time) < (now() - (%s || ' months')::interval))
                OR (MAX(a.punch_time) IS NULL
                    AND e.created_at < (now() - (%s || ' months')::interval))
            ORDER BY MAX(a.punch_time) NULLS FIRST""",
        tuple(params + [months, months]))
    return rows


def _purge_one(pin: str, devices: list) -> bool:
    if not db.query("SELECT 1 FROM employees WHERE pin=%s", (pin,)):
        return False
    for d in devices:
        db.execute(
            "INSERT INTO device_command (sn, content, kind, status) "
            "VALUES (%s,%s,'delete','pending')",
            (d["sn"], commands.delete_user_body(pin)))
    db.execute("DELETE FROM biotemplate WHERE pin=%s", (pin,))
    db.execute("DELETE FROM employees WHERE pin=%s", (pin,))
    return True


class PinsIn(BaseModel):
    pins: list[str]


@router.post("/purge-bulk")
def purge_bulk(body: PinsIn, user: dict = Depends(auth.require_admin)):
    """Delete many employees from ALL devices (queued) and the app DB — synchronized."""
    devices = db.query("SELECT sn FROM device")
    deleted = sum(1 for pin in body.pins if _purge_one(pin, devices))
    audit.log(user["id"], "purge_employees_bulk", "employee",
              f"{deleted} employees queued on {len(devices)} devices")
    return {"ok": True, "deleted": deleted, "devices_queued": len(devices)}


@router.post("/{pin}/purge")
def purge(pin: str, user: dict = Depends(auth.require_admin)):
    """Delete an employee from ALL devices (queued) and the app DB — synchronized."""
    devices = db.query("SELECT sn FROM device")
    if not _purge_one(pin, devices):
        raise HTTPException(404, "Employee not found")
    audit.log(user["id"], "purge_employee", "employee", f"{pin} (queued on {len(devices)} devices)")
    return {"ok": True, "pin": pin, "devices_queued": len(devices)}


PIN_START_DEFAULT = 1001
PIN_MAX_DEFAULT = 1999     # device capacity: IDs 1001..1999


def _setting_int(key: str, default: int) -> int:
    rows = db.query("SELECT value FROM settings WHERE key=%s", (key,))
    if rows and str(rows[0]["value"]).isdigit():
        return int(rows[0]["value"])
    return default


@router.get("/next-pin")
def next_pin(user: dict = Depends(auth.get_current_user)):
    """Lowest free numeric ID in [start..max] (default 1001..1999), filling gaps."""
    base = _setting_int("pin_start", PIN_START_DEFAULT)
    pin_max = _setting_int("pin_max", PIN_MAX_DEFAULT)
    used = db.query(
        "SELECT pin::bigint AS n FROM employees "
        "WHERE pin ~ '^[0-9]+$' AND pin::bigint >= %s ORDER BY n", (base,))
    expected = base
    for r in used:
        n = r["n"]
        if n == expected:
            expected += 1
        elif n > expected:
            break            # gap found at `expected`
    full = expected > pin_max
    return {"next_pin": None if full else str(expected),
            "capacity_full": full, "pin_max": pin_max}


@router.get("/{pin}/photo")
def employee_photo(pin: str, t: str = ""):
    """Serve an employee's photo. Auth via ?t=<token> so <img> tags can use it."""
    if not auth.decode_token(t):
        raise HTTPException(401, "Not authenticated")
    rows = db.query("SELECT photo FROM employees WHERE pin=%s", (pin,))
    if not rows or not rows[0]["photo"]:
        raise HTTPException(404, "No photo")
    data_uri = rows[0]["photo"]
    # stored as "data:image/jpeg;base64,...."
    try:
        header, b64 = data_uri.split(",", 1)
        mt = header.split(";")[0].replace("data:", "") or "image/jpeg"
        import base64
        return Response(base64.b64decode(b64), media_type=mt,
                        headers={"Cache-Control": "public, max-age=86400"})
    except Exception:
        raise HTTPException(500, "Bad photo data")


@router.get("/export.xlsx")
def export_xlsx(user: dict = Depends(auth.get_current_user),
                search: str | None = None, department_id: int | None = None):
    scope_sql, params = _scope_clause(user)
    where = ["1=1"]
    if search:
        where.append(
            "(e.pin ILIKE %s OR e.name ILIKE %s OR e.first_name ILIKE %s "
            "OR e.last_name ILIKE %s "
            "OR TRIM(COALESCE(e.first_name,'')||' '||COALESCE(e.last_name,'')) ILIKE %s)")
        s = f"%{search}%"
        params += [s, s, s, s, s]
    if department_id is not None:
        where.append("e.department_id = %s"); params.append(department_id)
    rows = db.query(f"{SELECT} WHERE {' AND '.join(where)} {scope_sql} ORDER BY e.pin", tuple(params))
    wb = Workbook(); ws = wb.active; ws.title = "Employees"
    headers = ["PIN", "Name", "Department", "Position", "Card", "Mobile", "Gender", "Hire date", "Active"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor="1e293b"); c.font = Font(color="FFFFFF", bold=True)
    for r in rows:
        ws.append([r["pin"], r["name"], r["department"] or "", r["position"] or "",
                   r["card"] or "", r["mobile"] or "", r["gender"] or "",
                   str(r["hire_date"] or ""), "yes" if r["active"] else "no"])
    buf = BytesIO(); wb.save(buf)
    return Response(buf.getvalue(), media_type=XLSX_MT, headers={
        "Content-Disposition": 'attachment; filename="employees.xlsx"'})


@router.post("/import")
async def import_xlsx(file: UploadFile = File(...), user: dict = Depends(auth.require_admin)):
    """Upsert employees from an .xlsx with columns PIN, Name, Department id (optional)."""
    wb = load_workbook(BytesIO(await file.read()), read_only=True)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None
    ci_pin, ci_name = col("pin", "emp_code"), col("name")
    ci_dept = col("department_id", "department id", "dept_id")
    ci_card = col("card")
    if ci_pin is None:
        raise HTTPException(400, "Missing 'PIN' column")
    created = updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[ci_pin] in (None, ""):
            continue
        pin = str(row[ci_pin]).strip()
        name = str(row[ci_name]).strip() if ci_name is not None and row[ci_name] else None
        dept = row[ci_dept] if ci_dept is not None else None
        card = str(row[ci_card]).strip() if ci_card is not None and row[ci_card] else None
        exists = db.query("SELECT 1 FROM employees WHERE pin=%s", (pin,))
        db.execute(
            """INSERT INTO employees (pin, name, department_id, card, updated_at)
               VALUES (%s,%s,%s,%s, now())
               ON CONFLICT (pin) DO UPDATE SET
                   name = COALESCE(EXCLUDED.name, employees.name),
                   department_id = COALESCE(EXCLUDED.department_id, employees.department_id),
                   card = COALESCE(EXCLUDED.card, employees.card), updated_at = now()""",
            (pin, name, int(dept) if dept else None, card))
        updated += 1 if exists else 0
        created += 0 if exists else 1
    return {"ok": True, "created": created, "updated": updated}


@router.get("/{pin}")
def get_employee(pin: str, user: dict = Depends(auth.get_current_user)):
    rows = db.query(f"{SELECT} WHERE e.pin = %s", (pin,))
    if not rows:
        raise HTTPException(404, "Employee not found")
    _assert_can_touch_dept(user, rows[0]["department_id"])
    return rows[0]


@router.post("")
def create_employee(body: EmployeeIn, user: dict = Depends(auth.get_current_user)):
    if not body.pin:
        raise HTTPException(400, "pin is required")
    _assert_can_touch_dept(user, body.department_id)
    if db.query("SELECT 1 FROM employees WHERE pin=%s", (body.pin,)):
        raise HTTPException(409, "pin already exists")
    name = body.name or " ".join(filter(None, [body.first_name, body.last_name])) or None
    db.execute(
        """INSERT INTO employees (pin, name, first_name, last_name, card,
                department_id, position_id, area_id, gender, mobile, hire_date, active)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, COALESCE(%s, TRUE))""",
        (body.pin, name, body.first_name, body.last_name, body.card,
         body.department_id, body.position_id, body.area_id, body.gender,
         body.mobile, body.hire_date or None, body.active))
    return {"ok": True, "pin": body.pin}


@router.put("/{pin}")
def update_employee(pin: str, body: EmployeeIn,
                    user: dict = Depends(auth.get_current_user)):
    cur = db.query("SELECT department_id FROM employees WHERE pin=%s", (pin,))
    if not cur:
        raise HTTPException(404, "Employee not found")
    _assert_can_touch_dept(user, cur[0]["department_id"])
    if body.department_id is not None:
        _assert_can_touch_dept(user, body.department_id)
    fields, params = [], []
    for col in ("name", "first_name", "last_name", "card", "department_id",
                "position_id", "area_id", "gender", "mobile", "active"):
        val = getattr(body, col)
        if val is not None:
            fields.append(f"{col}=%s")
            params.append(val)
    if body.hire_date is not None:
        fields.append("hire_date=%s")
        params.append(body.hire_date or None)
    if not fields:
        return {"ok": True}
    params.append(pin)
    db.execute(f"UPDATE employees SET {', '.join(fields)}, updated_at=now() "
               f"WHERE pin=%s", tuple(params))
    return {"ok": True}


@router.delete("/{pin}")
def delete_employee(pin: str, user: dict = Depends(auth.get_current_user)):
    cur = db.query("SELECT department_id FROM employees WHERE pin=%s", (pin,))
    if not cur:
        raise HTTPException(404, "Employee not found")
    _assert_can_touch_dept(user, cur[0]["department_id"])
    db.execute("UPDATE employees SET active=FALSE, updated_at=now() WHERE pin=%s", (pin,))
    return {"ok": True}
