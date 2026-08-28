"""Reports: daily/monthly summaries + Excel (xlsx) export."""
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from . import attendance_engine as eng
from . import auth, db
from .router_dashboard import dubai_today

router = APIRouter(prefix="/api/reports", tags=["reports"])

XLSX_MT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HDR_FILL = PatternFill("solid", fgColor="1e293b")
HDR_FONT = Font(color="FFFFFF", bold=True)


def _scope(user):
    ids = auth.scope_dept_ids(user)
    if ids is None:
        return "", []
    if not ids:
        return "AND FALSE", []
    return "AND e.department_id = ANY(%s)", [ids]


def _xlsx(title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HDR_FILL
        c.font = HDR_FONT
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, 1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = min(width + 3, 40)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _hm(m):
    m = m or 0
    return f"{m // 60}:{m % 60:02d}"


# ------------------------------ Daily xlsx ------------------------------ #
@router.get("/daily.xlsx")
def daily_xlsx(date: str, user: dict = Depends(auth.get_current_user)):
    d = datetime.strptime(date, "%Y-%m-%d").date()
    eng.recompute_if_stale(d)
    scope, p = _scope(user)
    rows = db.query(
        f"""SELECT ad.emp_code, {db.FULLNAME} AS name, dep.name AS department, ad.status,
               ad.first_in, ad.last_out, ad.worked_min, ad.late_min, ad.early_min,
               ad.ot_in, ad.ot_out, ad.ot_min
            FROM attendance_day ad
            LEFT JOIN employees e ON e.pin = ad.emp_code
            LEFT JOIN department dep ON dep.id = e.department_id
            WHERE ad.work_date = %s {scope} ORDER BY ad.emp_code""", tuple([d] + p))
    data = [[r["emp_code"], r["name"], r["department"] or "", r["status"],
             str(r["first_in"] or "")[11:19], str(r["last_out"] or "")[11:19],
             _hm(r["worked_min"]), r["late_min"], r["early_min"],
             str(r["ot_in"] or "")[11:19], str(r["ot_out"] or "")[11:19], _hm(r["ot_min"])]
            for r in rows]
    content = _xlsx(f"Daily {date}",
                    ["PIN", "Name", "Department", "Status", "Check-In", "Check-Out",
                     "Worked", "Late (m)", "Early (m)", "OT-In", "OT-Out", "OT"], data)
    return Response(content, media_type=XLSX_MT, headers={
        "Content-Disposition": f'attachment; filename="daily_{date}.xlsx"'})


# ------------------------------ Monthly ------------------------------ #
def _month_bounds(month: str):
    first = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    nxt = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    last = nxt - timedelta(days=1)
    return first, last


def _resolve_range(month, date_from, date_to):
    """Full month (YYYY-MM) OR an explicit date range (from/to). Range wins."""
    if date_from and date_to:
        first = datetime.strptime(date_from, "%Y-%m-%d").date()
        last = datetime.strptime(date_to, "%Y-%m-%d").date()
    else:
        first, last = _month_bounds(month)
    return first, last


def _monthly_rows(user: dict, month=None, date_from=None, date_to=None, include_absent=False):
    first, last = _resolve_range(month, date_from, date_to)
    end = min(last, dubai_today())
    d = first
    while d <= end:
        eng.recompute_if_stale(d)
        d += timedelta(days=1)
    scope, p = _scope(user)
    # Skip employees with no attendance in the range (long-time absent) unless asked.
    having = "" if include_absent else \
        "HAVING count(*) FILTER (WHERE ad.status IN ('present','halfday','incomplete')) > 0"
    return db.query(
        f"""SELECT ad.emp_code, {db.FULLNAME} AS name, dep.name AS department,
               count(*) FILTER (WHERE ad.status IN ('present','halfday','incomplete')) AS present_days,
               count(*) FILTER (WHERE ad.status='absent')  AS absent_days,
               count(*) FILTER (WHERE ad.status='halfday') AS halfday_days,
               count(*) FILTER (WHERE ad.late_min > 0)     AS late_days,
               count(*) FILTER (WHERE ad.early_min > 0)    AS early_days,
               COALESCE(sum(ad.worked_min),0) AS worked_min,
               COALESCE(sum(ad.ot_min),0)     AS ot_min
            FROM attendance_day ad
            LEFT JOIN employees e ON e.pin = ad.emp_code
            LEFT JOIN department dep ON dep.id = e.department_id
            WHERE ad.work_date BETWEEN %s AND %s {scope}
            GROUP BY ad.emp_code, e.pin, e.first_name, e.last_name, e.name, dep.name
            {having}
            ORDER BY ad.emp_code""",
        tuple([first, last] + p))


@router.get("/monthly")
def monthly(user: dict = Depends(auth.get_current_user), month: str | None = None,
            date_from: str | None = Query(None, alias="from"),
            date_to: str | None = Query(None, alias="to"),
            include_absent: bool = False):
    return _monthly_rows(user, month, date_from, date_to, include_absent)


@router.get("/monthly.xlsx")
def monthly_xlsx(user: dict = Depends(auth.get_current_user), month: str | None = None,
                 date_from: str | None = Query(None, alias="from"),
                 date_to: str | None = Query(None, alias="to"),
                 include_absent: bool = False):
    rows = _monthly_rows(user, month, date_from, date_to, include_absent)
    label = f"{date_from}_to_{date_to}" if (date_from and date_to) else month
    data = [[r["emp_code"], r["name"], r["department"] or "", r["present_days"],
             r["absent_days"], r["halfday_days"], r["late_days"], r["early_days"],
             _hm(r["worked_min"]), _hm(r["ot_min"])] for r in rows]
    content = _xlsx(f"Report {label}",
                    ["PIN", "Name", "Department", "Present", "Absent", "Half-day",
                     "Late days", "Early days", "Total Worked (h:m)", "Total OT (h:m)"], data)
    return Response(content, media_type=XLSX_MT, headers={
        "Content-Disposition": f'attachment; filename="report_{label}.xlsx"'})


# ---------------- Detailed per-employee daily report ---------------- #
def _detail_employees(user, month, date_from, date_to, search, include_absent):
    first, last = _resolve_range(month, date_from, date_to)
    end = min(last, dubai_today())
    d = first
    while d <= end:
        eng.recompute_if_stale(d)
        d += timedelta(days=1)

    scope_frag, sparams = _scope(user)
    where = f"1=1 {scope_frag}"
    params = [first, last] + list(sparams)
    if search:
        where += (" AND (e.pin ILIKE %s OR e.name ILIKE %s OR "
                  "TRIM(COALESCE(e.first_name,'')||' '||COALESCE(e.last_name,'')) ILIKE %s)")
        s = f"%{search}%"
        params += [s, s, s]
    if not include_absent:
        where += (" AND EXISTS (SELECT 1 FROM attendance_day a2 WHERE a2.emp_code=e.pin "
                  "AND a2.work_date BETWEEN %s AND %s "
                  "AND a2.status IN ('present','halfday','incomplete'))")
        params += [first, last]

    rows = db.query(
        f"""SELECT e.pin, {db.FULLNAME} AS name, dep.name AS department,
                   ad.work_date, ad.first_in, ad.last_out, ad.worked_min,
                   ad.late_min, ad.early_min, ad.ot_in, ad.ot_out, ad.ot_min, ad.status
            FROM employees e
            LEFT JOIN department dep ON dep.id = e.department_id
            JOIN attendance_day ad ON ad.emp_code = e.pin
                 AND ad.work_date BETWEEN %s AND %s
            WHERE {where}
            ORDER BY e.pin, ad.work_date""",
        tuple(params))

    # group by employee
    emps = {}
    for r in rows:
        e = emps.get(r["pin"])
        if not e:
            e = emps[r["pin"]] = {
                "pin": r["pin"], "name": r["name"], "department": r["department"],
                "days": [],
                "summary": {"present": 0, "absent": 0, "halfday": 0, "late": 0,
                            "early": 0, "worked_min": 0, "ot_min": 0},
            }
        e["days"].append({
            "date": str(r["work_date"]),
            "weekday": r["work_date"].strftime("%a"),
            "first_in": r["first_in"], "last_out": r["last_out"],
            "worked_min": r["worked_min"], "late_min": r["late_min"],
            "early_min": r["early_min"], "ot_in": r["ot_in"], "ot_out": r["ot_out"],
            "ot_min": r["ot_min"], "status": r["status"],
        })
        s = e["summary"]
        if r["status"] in ("present", "halfday", "incomplete"): s["present"] += 1
        if r["status"] == "absent": s["absent"] += 1
        if r["status"] == "halfday": s["halfday"] += 1
        if r["late_min"]: s["late"] += 1
        if r["early_min"]: s["early"] += 1
        s["worked_min"] += r["worked_min"] or 0
        s["ot_min"] += r["ot_min"] or 0
    return list(emps.values()), first, last


@router.get("/detail")
def detail(user: dict = Depends(auth.get_current_user), month: str | None = None,
           date_from: str | None = Query(None, alias="from"),
           date_to: str | None = Query(None, alias="to"),
           search: str | None = None, include_absent: bool = False):
    emps, first, last = _detail_employees(user, month, date_from, date_to, search, include_absent)
    return {"from": str(first), "to": str(last), "employees": emps}


@router.get("/detail.xlsx")
def detail_xlsx(user: dict = Depends(auth.get_current_user), month: str | None = None,
                date_from: str | None = Query(None, alias="from"),
                date_to: str | None = Query(None, alias="to"),
                search: str | None = None, include_absent: bool = False):
    emps, first, last = _detail_employees(user, month, date_from, date_to, search, include_absent)
    label = f"{first}_to_{last}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"[:31]
    hdr = ["Date", "Day", "Check-In", "Check-Out", "Worked", "Late (m)",
           "OT-In", "OT-Out", "OT", "Status"]
    row = 1
    for e in emps:
        # employee title
        c = ws.cell(row=row, column=1,
                    value=f"{e['pin']}  {e['name'] or ''}  —  {e['department'] or ''}")
        c.font = Font(bold=True, size=12)
        row += 1
        # column header
        for i, h in enumerate(hdr, 1):
            hc = ws.cell(row=row, column=i, value=h)
            hc.fill = HDR_FILL; hc.font = HDR_FONT
        row += 1
        for drow in e["days"]:
            vals = [drow["date"], drow["weekday"],
                    str(drow["first_in"] or "")[11:19], str(drow["last_out"] or "")[11:19],
                    _hm(drow["worked_min"]), drow["late_min"] or "",
                    str(drow["ot_in"] or "")[11:19], str(drow["ot_out"] or "")[11:19],
                    _hm(drow["ot_min"]) if drow["ot_min"] else "", drow["status"]]
            for i, v in enumerate(vals, 1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        sm = e["summary"]
        tc = ws.cell(row=row, column=1, value="TOTAL")
        tc.font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Present {sm['present']}  Absent {sm['absent']}")
        ws.cell(row=row, column=5, value=_hm(sm["worked_min"]))
        ws.cell(row=row, column=9, value=_hm(sm["ot_min"]))
        for cc in ws[row]:
            cc.font = Font(bold=True)
        row += 2  # blank spacer between employees
    for col, w in zip("ABCDEFGHIJ", (12, 6, 10, 10, 9, 8, 10, 10, 8, 12)):
        ws.column_dimensions[col].width = w
    buf = BytesIO(); wb.save(buf)
    return Response(buf.getvalue(), media_type=XLSX_MT, headers={
        "Content-Disposition": f'attachment; filename="attendance_{label}.xlsx"'})
